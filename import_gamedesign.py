"""
FitQuest -- Import Game Design Excel -> src/data/*.js
=====================================================
Usage :
    python import_gamedesign.py
    python import_gamedesign.py MonFichier.xlsx

Fichiers generes :
    src/data/zones.js
    src/data/bosses.js
    src/data/boss_mechanics.js   <- NOUVEAU
    src/data/exercises.js        <- NOUVEAU
    src/data/weapons.js
    src/data/spells.js
    src/data/recipes.js
    src/data/materials.js
    src/data/ingredients.js

REGLE IMPERATIVE :
    Toute modification du GameDesign (ajout de boss, equipement, mecanique...)
    doit toujours etre faite dans le fichier Excel ET refleter une mise a jour
    de ce script si necessaire. Les deux restent toujours synchronises.
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl manquant. Lance : pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / (sys.argv[1] if len(sys.argv) > 1 else "FitQuest_GameDesign.xlsx")
DATA_DIR   = SCRIPT_DIR / "fitquest" / "src" / "data"

# --- Definition des colonnes par feuille -------------------------------------

COLS = {
    "zones": [
        "id","name","element","levelMin","levelMax","requiredLevel",
        "requiredRegionalBoss","regionalBossId","travelKm","enterStepCost",
        "themeColor","accent","desc","svgKey"
    ],
    "boss": [
        "id","name","level","rarity","type","element","icon",
        "hp_max","attack","defense","gold","xp","desc","region","isRegionalBoss",
        "drop1_type","drop1_id","drop1_chance","drop1_qty_min","drop1_qty_max",
        "drop2_type","drop2_id","drop2_chance","drop2_qty_min","drop2_qty_max",
        "drop3_type","drop3_id","drop3_chance","drop3_qty_min","drop3_qty_max",
        "spawnRate","minPlayerLevel","maxPlayerLevel",
    ],
    "boss_mechanics": [
        "bossId","bossRarity",
        "mechanicId","mechanicName","family",
        "trigger","triggerHpPct","triggerTurnN",
        "param1Name","param1Value","param2Name","param2Value",
        "icon","playerMessage","notes",
    ],
    "exercises": [
        "id","name","icon","category",
        "muscleGroup","difficulty","isTimed","equipment",
        "xpPerRep","xpPerSecond","defaultReps","defaultDurationS",
        "bonusForce","bonusAgility","bonusConstitution","bonusEndurance",
        "desc","notes",
    ],
    "weapons": [
        "id","name","rarity","icon","slot","desc",
        "stats_force","stats_agility","stats_constitution","stats_defense"
    ],
    "spells": [
        "id","name","icon","element","manaCost","effect","value","oncePerCombat","desc"
    ],
    "recipes_blacksmith": [
        "weaponId","gold",
        "mat1_id","mat1_qty","mat2_id","mat2_qty","mat3_id","mat3_qty"
    ],
    "recipes_witch": [
        "id","name","icon","rarity","desc","effect","gold",
        "ing1_id","ing1_qty","ing2_id","ing2_qty"
    ],
    "materials":   ["id","name","rarity","icon","desc"],
    "ingredients": ["id","name","rarity","icon","desc"],
}

# --- Helpers -----------------------------------------------------------------

def _normalize(s):
    import unicodedata
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode().lower().strip()

def _valid_id(v):
    """ID valide : snake_case uniquement. Filtre les lignes de notes ou d'en-tetes."""
    if not v or not isinstance(v, str): return False
    import re as _re
    return bool(_re.match(r'^[a-z][a-z0-9_]*$', v.strip()))

def read_sheet(wb, key):
    sheet_map = {
        "zones":              "Zones",
        "boss":               "Boss",
        "boss_mechanics":     "Mecaniques Boss",
        "exercises":          "Exercices",
        "weapons":            "Armes",
        "spells":             "Sorts",
        "recipes_blacksmith": "Recettes Forgeron",
        "recipes_witch":      "Recettes Sorciere",
        "materials":          "Materiaux",
        "ingredients":        "Ingredients",
    }
    # Feuilles avec 2 lignes d'en-tete (section + colonnes) -> donnees a partir de row 3
    TWO_HEADER_ROWS = {"boss_mechanics", "exercises"}
    target = sheet_map[key]
    target_norm = _normalize(target)
    ws = None
    for name in wb.sheetnames:
        name_norm = _normalize(name)
        if target_norm in name_norm or name_norm == target_norm:
            ws = wb[name]
            break
    if ws is None:
        keywords = [w for w in target_norm.split() if len(w) > 3]
        for name in wb.sheetnames:
            name_norm = _normalize(name)
            if any(kw in name_norm for kw in keywords):
                ws = wb[name]
                break
    if ws is None:
        print("  [WARN] Onglet '" + target + "' introuvable -- ignore")
        return []
    cols = COLS[key]
    min_row = 3 if key in TWO_HEADER_ROWS else 2
    rows = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        first_val = row[0]
        if not _valid_id(str(first_val) if first_val else ""):
            continue  # saute les lignes de note, en-tete residuel, etc.
        d = {}
        for i, col_id in enumerate(cols):
            d[col_id] = row[i] if i < len(row) else None
        rows.append(d)
    return rows

def val(d, key, default=None):
    v = d.get(key, default)
    if v is None:
        return default
    if isinstance(v, str) and v.strip() == "":
        return default
    return v.strip() if isinstance(v, str) else v

def to_int(v, default=0):
    try:
        return int(v) if v not in (None, "") else default
    except:
        return default

def to_float(v, default=0.0):
    try:
        return float(v) if v not in (None, "") else default
    except:
        return default

def to_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().upper() in ("TRUE", "OUI", "YES", "1")
    return bool(v)

def js_val(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, str):
        return "'" + v.replace("\\", "\\\\").replace("'", "\\'") + "'"
    if isinstance(v, list):
        return "[" + ",".join(js_val(i) for i in v) + "]"
    if isinstance(v, dict):
        return "{" + ",".join(k + ":" + js_val(vv) for k, vv in v.items()) + "}"
    return js_val(str(v))

def obj_to_js(d):
    return "{" + ",".join(k + ":" + js_val(v) for k, v in d.items()) + "}"

def write_js(filepath, export_name, data):
    lines = ["export const " + export_name + " = ["]
    for i, item in enumerate(data):
        comma = "," if i < len(data) - 1 else ""
        lines.append("  " + obj_to_js(item) + comma)
    lines.append("];")
    filepath.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("  OK  " + str(filepath.relative_to(SCRIPT_DIR)) + "  (" + str(len(data)) + " entrees)")

def write_recipes_js(filepath, blacksmith, witch):
    lines = ["export const recipes_blacksmith = ["]
    for i, item in enumerate(blacksmith):
        comma = "," if i < len(blacksmith) - 1 else ""
        lines.append("  " + obj_to_js(item) + comma)
    lines.append("];\n")
    lines.append("export const recipes_witch = [")
    for i, item in enumerate(witch):
        comma = "," if i < len(witch) - 1 else ""
        lines.append("  " + obj_to_js(item) + comma)
    lines.append("];")
    filepath.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("  OK  " + str(filepath.relative_to(SCRIPT_DIR)))

# --- Parseurs ----------------------------------------------------------------

def parse_zones(rows):
    result = []
    for r in rows:
        zone_id = val(r, "id")
        if not zone_id:
            continue
        result.append({
            "id":                   zone_id,
            "name":                 val(r, "name", ""),
            "element":              val(r, "element", "wind"),
            "levelMin":             to_int(val(r, "levelMin")),
            "levelMax":             to_int(val(r, "levelMax")),
            "requiredLevel":        to_int(val(r, "requiredLevel")),
            "requiredRegionalBoss": val(r, "requiredRegionalBoss"),
            "regionalBossId":       val(r, "regionalBossId"),
            "enterStepCost":        to_int(val(r, "enterStepCost"), 0),
            "travelKm":             to_int(val(r, "travelKm"), 0),
            "themeColor":           val(r, "themeColor", "#6B7280"),
            "accent":               val(r, "accent", "#9CA3AF"),
            "desc":                 val(r, "desc", ""),
            "bgImage":              "/environnements/" + zone_id + "-dashboard.png",
            "combatBgImage":        "/environnements/" + zone_id + "-combat.png",
            "preSessionBgImage":    "/environnements/" + zone_id + "-presession.png",
            "svgKey":               val(r, "svgKey", zone_id),
        })
    return result

def parse_bosses(rows):
    result = []
    for r in rows:
        boss_id = val(r, "id")
        if not boss_id:
            continue
        drops = []
        for i in (1, 2, 3):
            prefix  = "drop" + str(i) + "_"
            dtype   = val(r, prefix + "type")
            did     = val(r, prefix + "id")
            dchance = to_float(val(r, prefix + "chance"), 0)
            dmin    = to_int(val(r, prefix + "qty_min"), 1)
            dmax    = to_int(val(r, prefix + "qty_max"), 1)
            if dtype and did and dchance > 0:
                drops.append({"type": dtype, "id": did, "chance": dchance, "qty": [dmin, dmax]})
        obj = {
            "id":             boss_id,
            "name":           val(r, "name", ""),
            "level":          to_int(val(r, "level"), 1),
            "rarity":         val(r, "rarity", "common"),
            "type":           val(r, "type", "force"),
            "element":        val(r, "element", "wind"),
            "icon":           val(r, "icon", "default"),
            "hp_max":         to_int(val(r, "hp_max"), 100),
            "attack":         to_int(val(r, "attack"), 10),
            "defense":        to_int(val(r, "defense"), 5),
            "gold":           to_int(val(r, "gold"), 10),
            "xp":             to_int(val(r, "xp"), 100),
            "desc":           val(r, "desc", ""),
            "region":         val(r, "region", "foret"),
            "spawnRate":      to_float(val(r, "spawnRate"), 0.5),
            "minPlayerLevel": to_int(val(r, "minPlayerLevel"), 1),
            "maxPlayerLevel": to_int(val(r, "maxPlayerLevel"), 99),
        }
        if to_bool(val(r, "isRegionalBoss", False)):
            obj["isRegionalBoss"] = True
        if drops:
            obj["drops"] = drops
        result.append(obj)
    return result

def parse_boss_mechanics(rows):
    by_boss = {}
    for r in rows:
        boss_id = val(r, "bossId")
        mec_id  = val(r, "mechanicId")
        if not boss_id or not mec_id or mec_id == "none":
            continue
        mec = {"id": mec_id}
        family = val(r, "family")
        if family and family not in ("--", None):
            mec["family"] = family
        trigger = val(r, "trigger")
        if trigger and trigger not in ("--", "always", None):
            mec["trigger"] = trigger
            hp_pct = to_float(val(r, "triggerHpPct"), 0)
            turn_n = to_int(val(r, "triggerTurnN"), 0)
            if hp_pct:
                mec["triggerHpPct"] = hp_pct
            if turn_n:
                mec["triggerTurnN"] = turn_n
        else:
            mec["trigger"] = "always"
        params = {}
        p1n = val(r, "param1Name")
        p1v = val(r, "param1Value")
        p2n = val(r, "param2Name")
        p2v = val(r, "param2Value")
        if p1n and p1v is not None and str(p1n) not in ("None", "--"):
            params[str(p1n)] = to_float(p1v) if isinstance(p1v, (int, float)) else p1v
        if p2n and p2v is not None and str(p2n) not in ("None", "--"):
            params[str(p2n)] = to_float(p2v) if isinstance(p2v, (int, float)) else p2v
        if params:
            mec["params"] = params
        icon = val(r, "icon")
        if icon and icon not in ("--", None):
            mec["icon"] = icon
        msg = val(r, "playerMessage")
        if msg and msg not in ("--", None):
            mec["playerMessage"] = msg
        if boss_id not in by_boss:
            by_boss[boss_id] = []
        by_boss[boss_id].append(mec)
    result = []
    for boss_id, mechanics in by_boss.items():
        result.append({"bossId": boss_id, "mechanics": mechanics})
    return result

def parse_exercises(rows):
    result = []
    for r in rows:
        ex_id = val(r, "id")
        if not ex_id:
            continue
        bonuses = {}
        for stat in ("Force", "Agility", "Constitution", "Endurance"):
            v = to_int(val(r, "bonus" + stat), 0)
            if v:
                bonuses[stat.lower()] = v
        obj = {
            "id":              ex_id,
            "name":            val(r, "name", ""),
            "icon":            val(r, "icon", "dumbbell"),
            "category":        val(r, "category", "force"),
            "muscleGroup":     val(r, "muscleGroup", "full_body"),
            "difficulty":      to_int(val(r, "difficulty"), 2),
            "isTimed":         to_bool(val(r, "isTimed", False)),
            "equipment":       val(r, "equipment", "aucun"),
            "xpPerRep":        to_int(val(r, "xpPerRep"), 0),
            "xpPerSecond":     to_int(val(r, "xpPerSecond"), 0),
            "defaultReps":     to_int(val(r, "defaultReps"), 0),
            "defaultDuration": to_int(val(r, "defaultDurationS"), 0),
            "desc":            val(r, "desc", ""),
        }
        if bonuses:
            obj["statBonuses"] = bonuses
        result.append(obj)
    return result

def parse_weapons(rows):
    result = []
    for r in rows:
        wid = val(r, "id")
        if not wid:
            continue
        stats = {}
        for stat in ("force", "agility", "constitution", "defense"):
            v = to_int(val(r, "stats_" + stat), 0)
            if v:
                stats[stat] = v
        result.append({
            "id":     wid,
            "name":   val(r, "name", ""),
            "rarity": val(r, "rarity", "common"),
            "icon":   val(r, "icon", "broadsword"),
            "stats":  stats,
            "slot":   val(r, "slot", "weapon_main"),
            "desc":   val(r, "desc", ""),
        })
    return result

def parse_spells(rows):
    result = []
    for r in rows:
        sid = val(r, "id")
        if not sid:
            continue
        result.append({
            "id":            sid,
            "name":          val(r, "name", ""),
            "icon":          val(r, "icon", "magic-swirl"),
            "element":       val(r, "element", "fire"),
            "manaCost":      to_int(val(r, "manaCost"), 10),
            "effect":        val(r, "effect", "damage_flat"),
            "value":         to_int(val(r, "value"), 0),
            "oncePerCombat": to_bool(val(r, "oncePerCombat", False)),
            "desc":          val(r, "desc", ""),
        })
    return result

def parse_recipes_blacksmith(rows):
    result = []
    for r in rows:
        wid = val(r, "weaponId")
        if not wid:
            continue
        mats = []
        for i in (1, 2, 3):
            mid  = val(r, "mat" + str(i) + "_id")
            mqty = to_int(val(r, "mat" + str(i) + "_qty"), 0)
            if mid and mqty > 0:
                mats.append({"id": mid, "qty": mqty})
        result.append({"weaponId": wid, "materials": mats, "gold": to_int(val(r, "gold"), 0)})
    return result

def parse_recipes_witch(rows):
    result = []
    for r in rows:
        rid = val(r, "id")
        if not rid:
            continue
        ings = []
        for i in (1, 2):
            iid  = val(r, "ing" + str(i) + "_id")
            iqty = to_int(val(r, "ing" + str(i) + "_qty"), 0)
            if iid and iqty > 0:
                ings.append({"id": iid, "qty": iqty})
        result.append({
            "id":          rid,
            "name":        val(r, "name", ""),
            "icon":        val(r, "icon", "standing-potion"),
            "rarity":      val(r, "rarity", "common"),
            "desc":        val(r, "desc", ""),
            "effect":      val(r, "effect", "heal"),
            "ingredients": ings,
            "gold":        to_int(val(r, "gold"), 0),
        })
    return result

def parse_simple(rows):
    result = []
    for r in rows:
        rid = val(r, "id")
        if not rid:
            continue
        result.append({
            "id":     rid,
            "name":   val(r, "name", ""),
            "rarity": val(r, "rarity", "common"),
            "icon":   val(r, "icon", "default"),
            "desc":   val(r, "desc", ""),
        })
    return result

# --- Main --------------------------------------------------------------------

def main():
    if not EXCEL_FILE.exists():
        print("Fichier introuvable : " + str(EXCEL_FILE))
        sys.exit(1)
    print("\nLecture de : " + EXCEL_FILE.name)
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("\nGeneration des fichiers JS dans " + str(DATA_DIR.relative_to(SCRIPT_DIR)) + " ...\n")

    rows = read_sheet(wb, "zones")
    if rows:
        write_js(DATA_DIR / "zones.js", "zones", parse_zones(rows))

    rows = read_sheet(wb, "boss")
    if rows:
        write_js(DATA_DIR / "bosses.js", "bosses", parse_bosses(rows))

    rows = read_sheet(wb, "boss_mechanics")
    if rows:
        write_js(DATA_DIR / "boss_mechanics.js", "boss_mechanics", parse_boss_mechanics(rows))

    rows = read_sheet(wb, "exercises")
    if rows:
        write_js(DATA_DIR / "exercises.js", "exercises", parse_exercises(rows))

    rows = read_sheet(wb, "weapons")
    if rows:
        write_js(DATA_DIR / "weapons.js", "weapons", parse_weapons(rows))

    rows = read_sheet(wb, "spells")
    if rows:
        write_js(DATA_DIR / "spells.js", "spells", parse_spells(rows))

    rows_rf = read_sheet(wb, "recipes_blacksmith")
    rows_rs = read_sheet(wb, "recipes_witch")
    if rows_rf or rows_rs:
        write_recipes_js(
            DATA_DIR / "recipes.js",
            parse_recipes_blacksmith(rows_rf),
            parse_recipes_witch(rows_rs),
        )

    rows = read_sheet(wb, "materials")
    if rows:
        write_js(DATA_DIR / "materials.js", "materials", parse_simple(rows))

    rows = read_sheet(wb, "ingredients")
    if rows:
        write_js(DATA_DIR / "ingredients.js", "ingredients", parse_simple(rows))

    print("\nImport termine ! Lance le jeu pour verifier les changements.\n")

if __name__ == "__main__":
    main()
